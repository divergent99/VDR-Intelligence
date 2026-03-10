# ingestion/extractor.py
"""
VDR document ingestion — extracts raw text from PDF, DOCX, and XLSX files.
Supports two modes:
  1. extract_from_folder()  — walk a local VDR folder path
  2. extract_from_upload()  — process a single in-memory file (FastAPI UploadFile bytes)
"""

from __future__ import annotations

import logging
import os
from pathlib import Path

from config import settings

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls"}


# ─────────────────────────────────────────────
# SINGLE FILE PARSERS
# ─────────────────────────────────────────────

def _parse_pdf(filepath: str) -> str:
    import fitz  # PyMuPDF
    doc = fitz.open(filepath)
    return "\n".join(page.get_text() for page in doc)


def _parse_docx(filepath: str) -> str:
    from docx import Document
    doc = Document(filepath)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_xlsx(filepath: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"\n[Sheet: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(v) for v in row if v is not None]
            if row_vals:
                lines.append(" | ".join(row_vals))
    return "\n".join(lines)


def _parse_file(filepath: str, ext: str) -> str:
    """Route a file to the correct parser by extension."""
    parsers = {
        ".pdf":  _parse_pdf,
        ".docx": _parse_docx,
        ".xlsx": _parse_xlsx,
        ".xls":  _parse_xlsx,
    }
    parser = parsers.get(ext)
    if not parser:
        raise ValueError(f"Unsupported file type: {ext}")
    return parser(filepath)


# ─────────────────────────────────────────────
# MODE 1 — FOLDER EXTRACTION
# ─────────────────────────────────────────────

def extract_from_folder(folder_path: str) -> str:
    """
    Walk a VDR folder, extract text from all supported files,
    concatenate into one string ready for the pipeline.

    Args:
        folder_path: Absolute path to the VDR folder.

    Returns:
        Combined extracted text, truncated to settings.doc_char_limit.

    Raises:
        FileNotFoundError: If the folder doesn't exist.
    """
    if not os.path.isdir(folder_path):
        raise FileNotFoundError(f"VDR folder not found: {folder_path}")

    all_text: list[str] = []

    for filename in sorted(os.listdir(folder_path)):
        ext = Path(filename).suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            continue

        filepath = os.path.join(folder_path, filename)
        logger.info("Ingesting: %s", filename)

        try:
            text = _parse_file(filepath, ext)
            all_text.append(
                f"\n\n{'='*60}\nDOCUMENT: {filename}\n{'='*60}\n{text}"
            )
        except Exception as exc:
            logger.warning("Failed to parse %s: %s", filename, exc)
            continue

    combined = "\n".join(all_text)

    if len(combined) > settings.doc_char_limit:
        logger.info(
            "Truncating combined text: %d → %d chars",
            len(combined), settings.doc_char_limit,
        )
        combined = combined[: settings.doc_char_limit]

    logger.info("Folder extraction complete: %d chars from %s", len(combined), folder_path)
    return combined


# ─────────────────────────────────────────────
# MODE 2 — IN-MEMORY UPLOAD EXTRACTION
# ─────────────────────────────────────────────

def extract_from_bytes(file_bytes: bytes, filename: str) -> str:
    """
    Extract text from an in-memory file (e.g. FastAPI UploadFile).

    Args:
        file_bytes: Raw file content as bytes.
        filename:   Original filename — used to determine parser.

    Returns:
        Extracted text string, truncated to settings.node_char_limit per file.

    Raises:
        ValueError: If the file type is not supported.
    """
    import tempfile

    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type '{ext}' for file '{filename}'")

    # Write to a temp file so existing parsers (fitz, docx, openpyxl) can open it
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        text = _parse_file(tmp_path, ext)
    finally:
        os.unlink(tmp_path)

    # Per-file cap so one giant file can't swamp the whole pipeline
    if len(text) > settings.node_char_limit:
        text = text[: settings.node_char_limit]

    logger.info("Extracted %d chars from uploaded file '%s'", len(text), filename)
    return text


def combine_uploads(texts: list[tuple[str, str]]) -> str:
    """
    Combine multiple extracted upload texts into one pipeline-ready string.

    Args:
        texts: List of (filename, extracted_text) tuples.

    Returns:
        Combined string, truncated to settings.doc_char_limit.
    """
    parts = [
        f"\n\n{'='*60}\nDOCUMENT: {filename}\n{'='*60}\n{text}"
        for filename, text in texts
    ]
    combined = "\n".join(parts)

    if len(combined) > settings.doc_char_limit:
        logger.info(
            "Truncating combined uploads: %d → %d chars",
            len(combined), settings.doc_char_limit,
        )
        combined = combined[: settings.doc_char_limit]

    return combined